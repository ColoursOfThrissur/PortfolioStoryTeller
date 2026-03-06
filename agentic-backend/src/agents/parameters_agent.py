"""
Parameters Agent - Step 1: Collect report parameters
"""
import os
from typing import Dict
from datetime import datetime, date
from dotenv import load_dotenv
import openpyxl
import pandas as pd
from pathlib import Path

load_dotenv()

from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage, SystemMessage


class ParametersAgent:
    """
    Step 1: Report Parameters
    
    Collects:
    - Client name
    - Reporting period (Q4-2025, custom dates, etc.)
    - Portfolio file (Excel/CSV)
    - Benchmark (default: S&P 500)
    - Optional: firm branding, account filters
    """
    
    def __init__(self):
        self.llm = ChatGoogleGenerativeAI(
            model="gemini-2.5-flash",
            temperature=0,
            google_api_key=os.getenv("GEMINI_API_KEY")
        )
    
    async def execute(self, collected_data: Dict) -> Dict:
        """
        Process and validate collected parameters
        
        Args:
            collected_data: {
                "client_name": "John Mitchell",
                "period": "Q4-2025" or {"start": "2025-10-01", "end": "2025-12-31"},
                "portfolio_file": "path/to/file.xlsx",
                "benchmark": "^GSPC" (optional),
                "accounts_filter": [...] (optional)
            }
        
        Returns:
            {
                "status": "complete",
                "parameters": {
                    "client_name": "John Mitchell",
                    "period": {"name": "Q4-2025", "start_date": "2025-10-01", "end_date": "2025-12-31"},
                    "benchmark": {"ticker": "^GSPC", "name": "S&P 500"},
                    "holdings": [...],
                    "accounts": [...]
                }
            }
        """
        
        try:
            print(f"[DEBUG] ParametersAgent received: {list(collected_data.keys())}")
            print(f"[DEBUG] Portfolio file path: {collected_data.get('portfolio_file')}")
            
            # Parse period
            period = await self._parse_period(collected_data.get("period", "Q4-2025"))
            
            # Parse portfolio file
            portfolio_data = await self._parse_portfolio_file(collected_data["portfolio_file"])
            
            # Set benchmark
            benchmark_ticker = collected_data.get("benchmark", os.getenv("DEFAULT_BENCHMARK", "^GSPC"))
            benchmark = await self._get_benchmark_info(benchmark_ticker)
            
            # Extract client name (from data or provided)
            client_name = collected_data.get("client_name") or portfolio_data.get("client_name", "Client")
            
            parameters = {
                "client_name": client_name,
                "period": period,
                "benchmark": benchmark,
                "holdings": portfolio_data["holdings"],
                "accounts": portfolio_data.get("accounts", []),
                "firm_name": os.getenv("FIRM_NAME", "Wealth Management Firm")
            }
            
            return {
                "status": "complete",
                "parameters": parameters,
                "message": f"Parameters set for {client_name} - {period['name']}"
            }
            
        except Exception as e:
            return {
                "status": "error",
                "error": str(e),
                "message": f"Failed to process parameters: {str(e)}"
            }
    
    async def _parse_period(self, period_input) -> Dict:
        """Parse period string or dict into standardized format"""
        
        if isinstance(period_input, dict):
            # Already has start/end dates
            return {
                "name": f"{period_input['start']} to {period_input['end']}",
                "start_date": period_input["start"],
                "end_date": period_input["end"]
            }
        
        # Parse period string (e.g., "Q4-2025", "2025", "YTD")
        period_str = str(period_input).upper()
        
        if "Q1" in period_str:
            year = period_str.split("-")[1] if "-" in period_str else str(datetime.now().year)
            return {
                "name": f"Q1-{year}",
                "start_date": f"{year}-01-01",
                "end_date": f"{year}-03-31"
            }
        elif "Q2" in period_str:
            year = period_str.split("-")[1] if "-" in period_str else str(datetime.now().year)
            return {
                "name": f"Q2-{year}",
                "start_date": f"{year}-04-01",
                "end_date": f"{year}-06-30"
            }
        elif "Q3" in period_str:
            year = period_str.split("-")[1] if "-" in period_str else str(datetime.now().year)
            return {
                "name": f"Q3-{year}",
                "start_date": f"{year}-07-01",
                "end_date": f"{year}-09-30"
            }
        elif "Q4" in period_str:
            year = period_str.split("-")[1] if "-" in period_str else str(datetime.now().year)
            return {
                "name": f"Q4-{year}",
                "start_date": f"{year}-10-01",
                "end_date": f"{year}-12-31"
            }
        elif "YTD" in period_str:
            year = str(datetime.now().year)
            return {
                "name": f"YTD-{year}",
                "start_date": f"{year}-01-01",
                "end_date": datetime.now().strftime("%Y-%m-%d")
            }
        else:
            # Default to Q4 of current year
            year = str(datetime.now().year)
            return {
                "name": f"Q4-{year}",
                "start_date": f"{year}-10-01",
                "end_date": f"{year}-12-31"
            }
    
    async def _parse_portfolio_file(self, file_path: str) -> Dict:
        """Parse Excel or CSV portfolio file"""
        
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"Portfolio file not found: {file_path}")
        
        if path.suffix in ['.xlsx', '.xls']:
            return self._parse_excel(path)
        elif path.suffix == '.csv':
            return self._parse_csv(path)
        else:
            raise ValueError(f"Unsupported file type: {path.suffix}")
    
    def _parse_excel(self, file_path: Path) -> Dict:
        """Parse Excel portfolio file"""
        
        wb = openpyxl.load_workbook(file_path)
        
        # Try to find Holdings sheet
        if "Holdings" in wb.sheetnames:
            ws = wb["Holdings"]
        else:
            ws = wb.active
        
        # Try to find Client_Info sheet
        client_name = None
        if "Client_Info" in wb.sheetnames:
            info_ws = wb["Client_Info"]
            for row in info_ws.iter_rows(min_row=2, max_row=2, values_only=True):
                if row[0]:
                    client_name = row[0]
                    break
        
        holdings = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            holding = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    if header.lower() in ['ticker', 'symbol']:
                        holding['ticker'] = row[i]
                    elif header.lower() in ['shares', 'quantity']:
                        holding['shares'] = float(row[i] or 0)
                    elif header.lower() in ['cost basis', 'cost_basis', 'basis']:
                        holding['cost_basis'] = float(row[i] or 0)
                    elif header.lower() in ['account', 'account name', 'account_name']:
                        holding['account'] = row[i]
            
            if 'ticker' in holding and 'shares' in holding:
                holdings.append(holding)
        
        return {
            "client_name": client_name,
            "holdings": holdings,
            "accounts": list(set([h.get('account', 'Main') for h in holdings]))
        }
    
    def _parse_csv(self, file_path: Path) -> Dict:
        """Parse CSV portfolio file"""
        
        df = pd.read_csv(file_path)
        
        holdings = []
        client_name = None
        
        for _, row in df.iterrows():
            holding = {}
            
            for col in df.columns:
                if col.lower() in ['ticker', 'symbol']:
                    holding['ticker'] = row[col]
                elif col.lower() in ['shares', 'quantity']:
                    holding['shares'] = float(row[col])
                elif col.lower() in ['cost basis', 'cost_basis', 'basis']:
                    holding['cost_basis'] = float(row[col])
                elif col.lower() in ['account', 'account name']:
                    holding['account'] = row[col]
                elif col.lower() in ['client', 'client name', 'client_name']:
                    if not client_name:
                        client_name = row[col]
            
            if 'ticker' in holding and 'shares' in holding:
                holdings.append(holding)
        
        return {
            "client_name": client_name,
            "holdings": holdings,
            "accounts": list(set([h.get('account', 'Main') for h in holdings]))
        }
    
    async def _get_benchmark_info(self, ticker: str) -> Dict:
        """Get benchmark name from ticker"""
        
        benchmark_names = {
            "^GSPC": "S&P 500",
            "^DJI": "Dow Jones Industrial Average",
            "^IXIC": "NASDAQ Composite",
            "^RUT": "Russell 2000",
            "SPY": "S&P 500 ETF",
            "QQQ": "NASDAQ 100 ETF",
            "AGG": "Bloomberg Aggregate Bond"
        }
        
        return {
            "ticker": ticker,
            "name": benchmark_names.get(ticker, ticker)
        }
